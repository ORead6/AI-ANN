import matplotlib.pyplot as mp
import numpy as np
from openpyxl import Workbook


def plotError():
    data = []
    with open("errorData.txt", "r") as file:
        for eachLine in file:
            if (eachLine != "0.0\n"):
                data.append(float(eachLine))

    lst = list(np.arange(1,len(data)+1))
    mp.plot(lst, data)
    mp.xlabel("Epochs")
    mp.ylabel("RMSE")
    mp.title("Error Over Time")
    mp.show()

def plotDot():
    x = []
    y = []
    with open("dotGraph.txt", "r") as file:
        for eachLine in file:
            if (eachLine != "0.0\n"):
                data = eachLine.split(",")
                x.append(data[0])
                y.append(data[1])

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "x"
    ws["B1"] = "y"
    for i in range(1, len(x)+1):
        ws['A%s'%str(i+1)] = float(x[i-1])
        ws['B%s'%str(i+1)] = float(y[i-1])

        wb.save("dotGraph.xlsx")


plotDot()